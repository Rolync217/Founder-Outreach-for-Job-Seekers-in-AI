"""
tools/export.py
Export pipeline database to a multi-sheet Excel file.

Usage:
    python tools/export.py                      # writes reports/outreach_export.xlsx
    python tools/export.py --out /path/to/file  # custom output path
"""

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

from tools.db_conn import get_conn


def _rows(sql: str, params: tuple = ()) -> list[dict]:
    with get_conn() as conn:
        return conn.execute(sql, params).fetchall()


def _json_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False)
    return str(val)


def export(output_path: str | None = None) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    Path("reports").mkdir(exist_ok=True)
    out = output_path or f"reports/outreach_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    wrap = Alignment(wrap_text=True, vertical="top")

    def _add_sheet(title: str, rows: list[dict]) -> None:
        ws = wb.create_sheet(title=title)
        if not rows:
            ws.append([f"No data yet — run the pipeline first."])
            return

        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in rows:
            ws.append([_json_to_str(row.get(h)) for h in headers])

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(header)),
                *(min(len(str(_json_to_str(r.get(header)) or "")), 60) for r in rows),
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 64)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap

        ws.freeze_panes = "A2"

    # ── Sheet 1: Companies ────────────────────────────────────────────────────
    companies = _rows("""
        SELECT
            c.name,
            c.company_website      AS website,
            c.sector,
            c.location,
            c.stage,
            c.team_size,
            c.founded_year,
            c.batch,
            c.source,
            c.status,
            c.description,
            c.created_at
        FROM public.companies_v2 c
        ORDER BY c.created_at DESC
    """)
    _add_sheet("Companies", companies)

    # ── Sheet 2: Scores ───────────────────────────────────────────────────────
    scores = _rows("""
        SELECT
            c.name                  AS company,
            c.company_website       AS website,
            s.total_score,
            s.tier,
            s.score_hiring          AS hiring,
            s.score_founder         AS founder,
            s.score_product         AS product,
            s.score_traction        AS traction,
            s.score_recency         AS recency,
            s.score_alignment       AS alignment,
            s.tier_reason,
            s.reasoning_hiring,
            s.reasoning_founder,
            s.reasoning_product,
            s.reasoning_traction,
            s.reasoning_recency,
            s.reasoning_alignment,
            s.created_at
        FROM public.scoring_v2 s
        JOIN public.companies_v2 c ON c.id = s.company_id
        ORDER BY s.tier ASC NULLS LAST, s.total_score DESC NULLS LAST
    """)
    _add_sheet("Scores", scores)

    # ── Sheet 3: Outreach Drafts ──────────────────────────────────────────────
    drafts = _rows("""
        SELECT
            c.name                  AS company,
            c.company_website       AS website,
            o.channel,
            o.message_mode,
            o.status,
            o.message_draft,
            o.message_final,
            o.opening,
            o.leverage_statement,
            o.relevant_experience,
            o.leverage_confidence,
            f.name                  AS founder_name,
            f.linkedin_url          AS founder_linkedin,
            o.created_at
        FROM public.outreach_v2 o
        JOIN public.companies_v2 c  ON c.id = o.company_id
        LEFT JOIN public.founders_v2 f ON f.id = o.founder_id
        WHERE o.is_deleted = false
        ORDER BY o.created_at DESC
    """)
    _add_sheet("Outreach Drafts", drafts)

    # ── Sheet 4: Research ─────────────────────────────────────────────────────
    research = _rows("""
        SELECT
            c.name                      AS company,
            c.company_website           AS website,
            r.problem_statement,
            r.solution_summary,
            r.product_type,
            r.has_live_product,
            r.has_paid_customers,
            r.is_agentic_ai,
            r.traction,
            r.tech_stack,
            r.alignment_notes,
            r.outreach_angle,
            r.funding_round,
            r.funding_amount,
            r.funding_investors,
            r.funding_date,
            r.hiring_signals_json       AS hiring_signals,
            r.company_founding_story,
            r.created_at
        FROM public.research_v2 r
        JOIN public.companies_v2 c ON c.id = r.company_id
        ORDER BY r.created_at DESC
    """)
    _add_sheet("Research", research)

    wb.save(out)
    return out


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export pipeline DB to Excel")
    parser.add_argument("--out", help="Output file path (default: reports/outreach_export_<timestamp>.xlsx)")
    args = parser.parse_args()

    path = export(output_path=args.out)
    print(f"Exported → {path}")
